import xlrd
import os
import sys
import openpyxl as op
import warnings
import time
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import range_boundaries
from openpyxl.chart.label import DataLabelList 
from openpyxl.chart import (
    LineChart,
    Reference,
    Series,
)
'''
	Script to take 5S scores for past X weeks and creates new 
	xlsx document with a tab for each zone, master tab, and
	graph for each zone with its respective data
'''

#Function that creates an array of all the columns being used
#Helps determine what the rightmost column is
#Returns the last column to the right
def Array_of_columns(filename, array):
	workbook = op.load_workbook(filename)
	ws = workbook.get_sheet_by_name('Sheet1')
	ws2 = workbook.get_sheet_by_name('Sheet2')

	num_of_cols = 0
	last_col = 0;
	for row in ws.iter_rows(min_col = 2, min_row =1, max_row =1): 
		for cell in row:
			#print('%s: cell.value=%s' % (cell, cell.value))
			array.append(cell.column)
			num_of_cols += 1;
			last_col = cell

	i = 0
	while i < len(columns):
		columns[i]+= '1'
		i += 1

	num_of_cols += 1;
	return num_of_cols;

#Copys the range of rows you would like
#Returns array of all te values
def copy_range(start_col, start_row, end_col, end_row, worksheet):
	range_selected = []
	#Loops through selected rows
	for i in range(start_row,end_row + 1,1):
		#Appends the row to a RowSelected list
		row_selected = []
		for j in range(start_col,end_col + 1,1):
			row_selected.append(ws.cell(row = i, column = j).value)
		#Adds the RowSelected list and nests it inside the rangeSelected
		range_selected.append(row_selected);
	return range_selected

#Paste data from copyRange into template sheet
def paste_range(start_col, start_row, end_col, end_row, sheetReceiving,copiedData):
    countRow = 0
    for i in range(start_row,end_row+1,1):
        countCol = 0
        for j in range(start_col,end_col+1,1):
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            sheetReceiving.cell(row = i, column = j).alignment = Alignment(horizontal='center')
            countCol += 1
        countRow += 1

''' Function that copy and pastes the data in one go instead of having to
	run two functions
'''
def createData(wb,new_ws,start_col1,start_row1,end_col1,end_row1,start_col2,start_row2,end_col2,end_row2):
	print("Copying and Pasting data...")
	selectedRange = copy_range(start_col1,start_row1,end_col1,end_row1,ws)
	pastingRange = paste_range(start_col2,start_row2,end_col2,end_row2,new_ws,selectedRange)
	workbook.save(wb_name)
	print("Range copied and pasted!")

#Function to create a chart for each individual zone
def create_chart(workbook,worksheet,zone,min_col,min_row,max_col,max_row):
	chart = LineChart()
	#Sets titles on graph
	chart.y_axis.title = '5S Audit Socres'
	chart.x_axis.title = '5S Weekly Audit Dates'
	chart.title = 'Past '+ string_weeks + ' Weeks Scores For Zone '+ zone

	#Gets the data from the table
	data = Reference(worksheet, min_col=min_col, min_row=min_row+1, max_col=max_col, max_row=max_row)

	#Gets the Dates from the first row and stores them in a variable
	dates = Reference(worksheet,min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row-1)

	#Creates a series using the data collected from the specified rows of the table
	series = Series(data,title = 'Zone ' + zone);
	#Sets the marker on the chart to be a triangle
	series.marker.symbol = "triangle"

	#Adds the data and date labels to the chart
	chart.append(series)
	chart.set_categories(dates)
	chart.dataLabels = DataLabelList()
	chart.dataLabels.showVal = True

	#Adds the chart to the worksheet at cell A10 and saves the file
	worksheet.add_chart(chart,'A10');
	workbook.save(wb_name);

#Creates Chart for the All Zones tab
def create_charts_multiple(worksheet):
	chart = LineChart();
	#Sets titles on graph
	chart.y_axis.title = '5S Audit Socres'
	chart.x_axis.title = '5S Weekly Audit Dates'
	chart.title = 'Past '+ string_weeks + ' Weeks Scores For All Zones'

	for i in range(1,13):
		#Gets the data from the table
		data = Reference(worksheet, min_col=2, min_row=1+i, max_col=weeks+1, max_row=1+i)
		#Gets the Dates from the first row and stores them in a variable
		dates = Reference(worksheet,min_col=2, min_row=1, max_col=weeks+1, max_row=1)
		#Creates a series using the data collected from the specified rows of the table
		series = Series(data,title = 'Zone ' + str(i));
		#Adds the data and date labels to the chart
		chart.append(series)
		chart.set_categories(dates)
		#chart.dataLabels = DataLabelList()
		#chart.dataLabels.showVal = True

	#Adds the chart to the worksheet at cell A10 and saves the file
	chart.height = 10;
	chart.width = 20;
	worksheet.add_chart(chart,'A18');
	workbook.save(wb_name);

#Function to create a chart for the average of the building
def create_avg_chart(workbook,worksheet,min_col,min_row,max_col,max_row):
	chart = LineChart()
	#Sets titles on graph
	chart.y_axis.title = '5S Audit Socre Average'
	chart.x_axis.title = '5S Weekly Audit Dates'
	chart.title = 'Past '+ string_weeks + ' Weeks Average Score For 539'

	#Gets the data from the table
	data = Reference(worksheet, min_col=min_col, min_row=min_row+14, max_col=max_col, max_row=max_row)

	#Gets the Dates from the first row and stores them in a variable
	dates = Reference(worksheet,min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row-14)

	#Creates a series using the data collected from the specified rows of the table
	series = Series(data,title = 'Building 539');
	#Sets the marker on the chart to be a triangle
	series.marker.symbol = "triangle"

	#Adds the data and date labels to the chart
	chart.append(series)
	chart.set_categories(dates)
	chart.dataLabels = DataLabelList()
	chart.dataLabels.showVal = True

	#Adds the chart to the worksheet at cell M18 and saves the file
	worksheet.add_chart(chart,'K18');
	workbook.save(wb_name);

#Finds the average For each zone and pastes it 2 coulmns to the right
#returns the average
def find_avg(worksheet,start_col,start_row,end_col,end_row):
	total  = 0
	average = 0
	counter = 0
	for i in range(start_row,end_row + 1,1):
		for j in range(start_col,end_col + 1,1):
			if worksheet.cell(row = i, column = j).value == None:
				total += 0.0
			else:
				total += float(worksheet.cell(row = i, column = j).value)
				counter +=1
	average = total/counter
	return average 


##############################################################################################################


#Gets input from user to figure out how many weeks back they would like to look at
correct = False
while(correct == False):	
	weeks = input("Only type numbers ex. 0123456789\nHow many weeks back would you like? ")
	valid = set('0123456789');
	if any((c in valid) for c in weeks):
		correct = True
	else:
		correct = False
		print("\nError: Only enter Numbers\n");

#Converts the number of weeks to an int value
string_weeks = weeks
weeks = int(weeks);

columns = [];

#Opens the master excel sheet
workbook = op.load_workbook('test.xlsx')
ws = workbook.get_sheet_by_name('Sheet1')
ws2 = workbook.get_sheet_by_name('Sheet2')

#Creates the title for the new Document with todays date
date = date.today()
wb_name = "5S Audit Scores " + str(date.strftime("%m.%d.%y")) +".xlsx"

#Creates the new workbook with an all zone sheet and a sheet for each zone
new_workbook = Workbook()
new_workbook.create_sheet("All Zones")
for i in range(1,13):
	new_workbook.create_sheet("Zone " + str(i))

#Removes the original default sheet and saves the excel file
new_ws = new_workbook.get_sheet_by_name('Sheet')
new_workbook.remove_sheet(new_ws)
new_workbook.save(wb_name)

#Figures out how many weeks of data there are in total
num_of_cols = Array_of_columns('test.xlsx',columns)


#Creates A sheet with all the zones
new_ws = new_workbook.get_sheet_by_name("All Zones");
#Adds the Zone Column (Not dynamic its based on only 12 zones)
createData(new_workbook,new_ws,1,2,1,13,1,2,1,13);
#Adds the date row (Not dynamic its based on only 12 zones)
createData(new_workbook,new_ws,num_of_cols-weeks+1,1,num_of_cols,1,2,1,weeks+1,1);
#Adds the rows scores for each zone and the avg (Not dynamic its based on only 12 zones)
createData(new_workbook,new_ws,num_of_cols-weeks+1,2,num_of_cols+1,13,2,2,weeks+1,13);
#Creates a chart that has all of the zones in one (Not dynamic its based on only 12 zones)
create_charts_multiple(new_ws)

#Loops through to add data to each zone tab
for i in range(1,13):
	#Gets the correct zone starting at 1
	zone_row = i+1;

	#Gets the zone sheet by name
	new_ws = new_workbook.get_sheet_by_name('Zone ' + str(i));

	#Adds zone column
	createData(new_workbook,new_ws,1,zone_row,1,zone_row,1,2,1,2);
	#Adds the row of dates
	createData(new_workbook,new_ws,num_of_cols-weeks+1,1,num_of_cols,1,2,1,weeks+1,1);
	#Adds the scores for the corresponding zone
	createData(new_workbook,new_ws,num_of_cols-weeks+1,zone_row,num_of_cols+1,zone_row,2,2,weeks+1,2);
	#Creates a chart for the data
	create_chart(new_workbook,new_ws,str(i),2,1,weeks+1,2);

#Creates a new chart for the average of the building
new_ws = new_workbook.get_sheet_by_name("All Zones");

#Creates title for coumn that has the avg scores for each zone
new_ws.cell(row = 1, column = weeks+3).value = 'Average Score For Each Zone'
#Sets alignment for the column to be centered
new_ws.cell(row = 1, column = weeks+3).alignment = Alignment(horizontal='center')

#Loop to calculate each zone avg and paste it in the column
for i in range(1,13):
	avg = find_avg(new_ws,2,i+1,weeks+1,i+1)
	new_ws.cell(row = i+1, column = weeks+3).value = round(avg,3)
	new_ws.cell(row = i+1, column = weeks+3).alignment = Alignment(horizontal='center')

#Loop that calulates the building average for the week
for i in range(1,weeks+1):
	avg = find_avg(new_ws,i+1,2,i+1,13)
	new_ws.cell(row = 15, column = i+1 ).value = round(avg,3)
	new_ws.cell(row = 15, column = i+1 ).alignment = Alignment(horizontal='center')
	new_workbook.save(wb_name)

#Sets and aligns the title cell for the building avg row
new_ws.cell(row = 15,column = 1).value = '539 Building Average'
new_ws.cell(row = 15,column = 1).alignment = Alignment(horizontal='center')

#Creates a chart for the building avg
create_avg_chart(new_workbook,new_ws,2,1,weeks+1,15)

#sets the width of the zone column and zone avg
new_ws.column_dimensions['A'].width = 19
new_ws.column_dimensions[new_ws.cell(row = 1, column=weeks+3).column].width = 25

new_workbook.save(wb_name)


